from flask import Blueprint, render_template, request, redirect, url_for, flash, session
from app import db
from app.models import Usuario, Solicitacao
from sqlalchemy.exc import IntegrityError
from datetime import datetime, date
from flask import json
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from flask import send_file


print("--- ROTAS CARREGADAS COM SUCESSO ---")

bp = Blueprint('main', __name__)
@bp.app_template_filter('datetimeformat')
def datetimeformat(value, format='%d-%m-%y'):
    try:
        # tenta converter string do tipo "2025-12-09"
        return datetime.strptime(value, "%Y-%m-%d").strftime(format)
    except:
        return value  # se falhar, retorna como está

# --- Context Processor: Simula o 'current_user' para o HTML ---
@bp.context_processor
def inject_user():
    class MockUser:
        is_authenticated = 'user_id' in session
        name = session.get('user_nome')
        id = session.get('user_id')
        tipo_usuario = session.get('user_tipo')
    return dict(current_user=MockUser())

# --- DASHBOARD UVIS ---

@bp.route('/')
def dashboard():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    # AJUSTE CHAVE: Se for admin, operario OU visualizar, redireciona para o painel de gestão
    if session.get('user_tipo') in ['admin', 'operario', 'visualizar']:
        return redirect(url_for('main.admin_dashboard'))

    try:
        user_id = int(session.get('user_id'))
    except (ValueError, TypeError):
        session.clear()
        flash('Sessão Inválida. Por favor, faça login novamente.', 'warning')
        return redirect(url_for('main.login'))

    # 1. Query Base: Pega os pedidos SÓ deste usuário
    query = Solicitacao.query.filter_by(usuario_id=user_id)

    # 2. Lógica do Filtro: Verifica se veio algo na URL (ex: ?status=PENDENTE)
    filtro_status = request.args.get('status')

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    # 3. Lógica da Paginação:
    page = request.args.get("page", 1, type=int)

    paginacao = query.order_by(
        Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6, error_out=False)

    return render_template(
        'dashboard.html',
        nome=session.get('user_nome'),
        solicitacoes=paginacao.items,
        paginacao=paginacao
    )

# --- PAINEL DE GESTÃO (Visualização para todos) ---
@bp.route('/admin')
def admin_dashboard():
    # AJUSTE CHAVE: Permite 'admin', 'operario' E 'visualizar'
    if 'user_id' not in session or session.get('user_tipo') not in ['admin', 'operario', 'visualizar']:
        flash('Acesso restrito.', 'danger')
        return redirect(url_for('main.login'))
    
    # Flag para controlar a renderização dos botões de edição no template
    is_editable = session.get('user_tipo') in ['admin', 'operario']
    
    # --- Captura filtros enviados pelo GET ---
    filtro_status = request.args.get("status")
    filtro_unidade = request.args.get("unidade")
    filtro_regiao = request.args.get("regiao")

    # --- Query base ---
    query = Solicitacao.query.join(Usuario)

    # --- Filtros aplicáveis ---
    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    page = request.args.get("page", 1, type=int)

    paginacao = query.order_by(
    Solicitacao.data_criacao.desc()
    ).paginate(page=page, per_page=6)

    return render_template(
    'admin.html',
    pedidos=paginacao.items,
    paginacao=paginacao,
    is_editable=is_editable # Variável enviada ao template para controle de formulário
)

@bp.route('/admin/exportar_excel')
def exportar_excel():
    # AJUSTE CHAVE: Permite APENAS 'admin' E 'operario'
    if 'user_id' not in session or session.get('user_tipo') not in ['admin', 'operario']:
        flash('Permissão negada para exportar.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    # --- Captura filtros ---
    filtro_status = request.args.get("status")
    filtro_unidade = request.args.get("unidade")
    filtro_regiao = request.args.get("regiao")

    # Query base
    query = Solicitacao.query.join(Usuario)

    if filtro_status:
        query = query.filter(Solicitacao.status == filtro_status)

    if filtro_unidade:
        query = query.filter(Usuario.nome_uvis.ilike(f"%{filtro_unidade}%"))

    if filtro_regiao:
        query = query.filter(Usuario.regiao.ilike(f"%{filtro_regiao}%"))

    pedidos = query.order_by(Solicitacao.data_criacao.desc()).all()

    # --- CRIA EXCEL ---
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório de Solicitações"

    # Cabeçalho ATUALIZADO (do segundo código, com mais campos)
    headers = [
        "ID", "Unidade", "Região",
        "Data Agendada", "Hora",
        "CEP", "Logradouro", "Número", "Bairro", "Cidade/UF", "Complemento",
        "Latitude", "Longitude",
        "Foco", "Tipo Visita", "Altura", "Criadouro?", "Apoio CET?",
        "Observação",
        "Status", "Protocolo", "Justificativa"
    ]

    # Estilos
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Escreve cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    # Conteúdo
    row_num = 2
    for p in pedidos:
        # Tratamento de Endereço (baseado no segundo código, mas mantendo a UF separada para clareza)
        cidade_uf = f"{p.cidade or ''}/{p.uf or ''}"
        logradouro_num = f"{p.logradouro or ''}"

        # Tratamento de Booleans (Sim/Não) do segundo código
        criadouro_txt = "SIM" if getattr(p, 'criadouro', None) else "NÃO"
        cet_txt = "SIM" if getattr(p, 'apoio_cet', None) else "NÃO"

        # Formatação de data (Corrigido o erro de importação de datetime)
        if p.data_agendamento:
            try:
                if isinstance(p.data_agendamento, (date, datetime)): 
                    data_formatada = p.data_agendamento.strftime("%d-%m-%y")
                # Se for string (caso do primeiro código)
                else:
                    data_formatada = datetime.strptime(str(p.data_agendamento), "%Y-%m-%d").strftime("%d-%m-%y")
            except ValueError:
                data_formatada = str(p.data_agendamento)
        else:
            data_formatada = ""

        row = [
            p.id,
            p.autor.nome_uvis,
            p.autor.regiao,
            data_formatada,
            p.hora_agendamento,
            p.cep,
            logradouro_num,
            getattr(p, 'numero', ''),
            p.bairro,
            cidade_uf,
            getattr(p, 'complemento', ''),
            getattr(p, 'latitude', ''),
            getattr(p, 'longitude', ''),
            p.foco,
            getattr(p, 'tipo_visita', ''),
            getattr(p, 'altura_voo', ''),
            criadouro_txt,
            cet_txt,
            getattr(p, 'observacao', ''),
            p.status,
            p.protocolo,
            p.justificativa
        ]

        for col_num, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

        row_num += 1

    # Freeze Pane (Mantido do primeiro código)
    ws.freeze_panes = "A2"

    # Ajuste automático de largura (Lógica do primeiro código, mas com a correção de 'column' para 'column_letter')
    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter 

        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        adjusted_width = max_length + 2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Enviar arquivo
    return send_file(
        output,
        download_name="relatorio_solicitacoes.xlsx",
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# --- ROTA DE ATUALIZAÇÃO ---
@bp.route('/admin/atualizar/<int:id>', methods=['POST'])
def atualizar(id):
    # AJUSTE CHAVE: Permite APENAS 'admin' E 'operario'
    if session.get('user_tipo') not in ['admin', 'operario']:
        flash('Permissão negada para esta ação.', 'danger')
        return redirect(url_for('main.admin_dashboard'))

    pedido = Solicitacao.query.get_or_404(id)

    # Campos de Geo do segundo código:
    pedido.protocolo = request.form.get('protocolo')
    pedido.status = request.form.get('status')
    pedido.justificativa = request.form.get('justificativa')
    pedido.latitude = request.form.get('latitude')
    pedido.longitude = request.form.get('longitude')


    db.session.commit()
    flash('Pedido atualizado com sucesso!', 'success')

    return redirect(url_for('main.admin_dashboard'))

# --- NOVO PEDIDO ---
@bp.route('/novo_cadastro', methods=['GET', 'POST'], endpoint='novo')
def novo():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    hoje = date.today().isoformat()

    if request.method == 'POST':
        try:
            user_id_int = int(session['user_id'])

            data_str = request.form.get('data')
            hora_str = request.form.get('hora')

            if data_str:
                data_obj = datetime.strptime(data_str, '%Y-%m-%d').date()
            else:
                data_obj = None

            if hora_str:
                hora_obj = datetime.strptime(hora_str, '%H:%M').time()
            else:
                hora_obj = None

            criadouro_bool = request.form.get('criadouro') == 'sim'
            apoio_cet_bool = request.form.get('apoio_cet') == 'sim'


            nova_solicitacao = Solicitacao(
                data_agendamento=data_obj,
                hora_agendamento=hora_obj,

                cep=request.form.get('cep'),
                logradouro=request.form.get('logradouro'),
                bairro=request.form.get('bairro'),
                cidade=request.form.get('cidade'),
                numero=request.form.get('numero'),
                uf=request.form.get('uf'),
                complemento=request.form.get('complemento'), 

                foco=request.form.get('foco'),

                tipo_visita=request.form.get('tipo_visita'),
                altura_voo=request.form.get('altura_voo'),
                criadouro=criadouro_bool,
                apoio_cet=apoio_cet_bool,
                observacao=request.form.get('observacao'),

                latitude=request.form.get('latitude'),
                longitude=request.form.get('longitude'),

                usuario_id=user_id_int,
                status='PENDENTE'
            )

            db.session.add(nova_solicitacao)
            db.session.commit()

            flash('Pedido enviado!', 'success')
            return redirect(url_for('main.dashboard'))

        except ValueError as ve:
            db.session.rollback()
            flash(f"Erro no formato de data/hora: {ve}", "warning")
        except Exception as e:
            db.session.rollback()
            flash(f"Erro ao salvar: {e}", "danger")

    return render_template('cadastro.html', hoje=hoje)

# --- LOGIN ---
@bp.route('/login', methods=['GET', 'POST'])
def login():
    if 'user_id' in session:
        # AJUSTE CHAVE: Redireciona para admin_dashboard se for admin, operario OU visualizar
        if session.get('user_tipo') in ['admin', 'operario', 'visualizar']:
            return redirect(url_for('main.admin_dashboard'))
        return redirect(url_for('main.dashboard'))

    if request.method == 'POST':
        user = Usuario.query.filter_by(login=request.form.get('login')).first()

        if user and user.check_senha(request.form.get('senha')):
            session['user_id'] = int(user.id)
            session['user_nome'] = user.nome_uvis
            session['user_tipo'] = user.tipo_usuario

            flash(f'Bem-vindo, {user.nome_uvis}! Login realizado com sucesso.', 'success')

            # AJUSTE CHAVE: Redireciona para admin_dashboard se for admin, operario OU visualizar
            if user.tipo_usuario in ['admin', 'operario', 'visualizar']:
                return redirect(url_for('main.admin_dashboard'))
            return redirect(url_for('main.dashboard'))
        else:
            flash('Login ou senha incorretos. Tente novamente.', 'danger')

    return render_template('login.html')

# --- LOGOUT ---
@bp.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('main.login'))

@bp.route("/forcar_erro")
def forcar_erro():
    1 / 0  # erro proposital
    return "nunca vai chegar aqui"

# --- RELATORIOS (Acesso Exclusivo Admin) ---
@bp.route('/relatorios')
def relatorios():
    if 'user_id' not in session:
        return redirect(url_for('main.login'))

    # VERIFICAÇÃO CHAVE: APENAS 'admin' tem acesso
    if session.get('user_tipo') != 'admin':
        flash('Acesso restrito aos administradores.', 'danger')
        # Redireciona para o painel de gestão, que o 'operario' e 'visualizar' podem ver
        return redirect(url_for('main.admin_dashboard')) 

    # --- CÓDIGO DE RELATÓRIOS (Com Novas Métricas) ---
    mes_atual = int(request.args.get('mes', datetime.now().month))
    ano_atual = int(request.args.get('ano', datetime.now().year))

    query_base = Solicitacao.query

    filtro_data = f'{ano_atual}-{mes_atual:02d}'

    query_filtrada = query_base.filter(
        db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data
    )

    # Métricas Principais
    total_solicitacoes = query_filtrada.count()
    total_aprovadas = query_filtrada.filter_by(status='APROVADO').count()
    total_recusadas = query_filtrada.filter_by(status='NEGADO').count()
    total_analise = query_filtrada.filter_by(status='EM ANÁLISE').count()
    total_pendentes = query_filtrada.filter_by(status='PENDENTE').count()

    # Dados por Região (Mantido)
    dados_regiao_raw = (
        db.session.query(Usuario.regiao, db.func.count(Solicitacao.id))
        .join(Usuario, Usuario.id == Solicitacao.usuario_id)
        .filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
        .group_by(Usuario.regiao)
        .all()
    )
    dados_regiao = [tuple(row) for row in dados_regiao_raw]
    
    # ----------------------
    # NOVAS MÉTRICAS DETALHADAS (Filtradas por Mês/Ano)
    # ----------------------
    
    # Detalhamento de Solicitações por Status (Para incluir 'PENDENTE'/'EM ANÁLISE' de forma flexível)
    dados_status_raw = (
        db.session.query(Solicitacao.status, db.func.count(Solicitacao.id))
        .filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
        .group_by(Solicitacao.status)
        .all()
    )
    dados_status = [tuple(row) for row in dados_status_raw]

    # Solicitações por Foco
    dados_foco_raw = (
        db.session.query(Solicitacao.foco, db.func.count(Solicitacao.id))
        .filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
        .group_by(Solicitacao.foco)
        .all()
    )
    dados_foco = [tuple(row) for row in dados_foco_raw]
    
    # Solicitações por Tipo de Visita
    dados_tipo_visita_raw = (
        db.session.query(Solicitacao.tipo_visita, db.func.count(Solicitacao.id))
        .filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
        .group_by(Solicitacao.tipo_visita)
        .all()
    )
    dados_tipo_visita = [tuple(row) for row in dados_tipo_visita_raw]

    # Solicitações por Altura de Voo
    dados_altura_voo_raw = (
        db.session.query(Solicitacao.altura_voo, db.func.count(Solicitacao.id))
        .filter(db.func.strftime('%Y-%m', Solicitacao.data_criacao) == filtro_data)
        .group_by(Solicitacao.altura_voo)
        .order_by(Solicitacao.altura_voo)
        .all()
    )
    dados_altura_voo = [tuple(row) for row in dados_altura_voo_raw]
    
    # Dados Mensais (Mantido)
    dados_mensais_raw = (
        db.session.query(
            db.func.strftime('%Y-%m', Solicitacao.data_criacao).label('mes'),
            db.func.count(Solicitacao.id)
        )
        .group_by('mes')
        .order_by('mes')
        .all()
    )
    dados_mensais = [tuple(row) for row in dados_mensais_raw]

    anos_disponiveis = sorted(list(set([d[0].split('-')[0] for d in dados_mensais])), reverse=True)

    return render_template(
        'relatorios.html',
        total_solicitacoes=total_solicitacoes,
        total_aprovadas=total_aprovadas,
        total_recusadas=total_recusadas,
        total_analise=total_analise,
        total_pendentes=total_pendentes,
        dados_regiao=dados_regiao,
        dados_mensais=dados_mensais,
        
        # NOVOS DADOS
        dados_status=dados_status,
        dados_foco=dados_foco,
        dados_tipo_visita=dados_tipo_visita,
        dados_altura_voo=dados_altura_voo,

        mes_selecionado=mes_atual,
        ano_selecionado=ano_atual,
        anos_disponiveis=anos_disponiveis
    )